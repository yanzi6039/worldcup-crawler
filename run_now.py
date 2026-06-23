"""
世界杯新闻爬虫 · CLI 入口

用法：
  python3 run_now.py --quick           # 快速测试：每源 5 篇
  python3 run_now.py --backfill        # 存量：每源 20 篇
  python3 run_now.py --incremental     # 增量：跳过已有，每源 30 篇
  python3 run_now.py --source=fifa,espn
  python3 run_now.py --serve           # 启动 Web + 后台定时爬虫（24/7）
  python3 run_now.py --init-db         # 重建数据库
"""
import os
import sys
import argparse
import time
import logging

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

# 强制北京时区（防止 cron 启动时 TZ=UTC 导致日期错位 + subprocess push 失败）
os.environ["TZ"] = "Asia/Shanghai"
try:
    time.tzset()
except AttributeError:
    pass  # Windows 无 tzset

from config import LOG_DIR, WEB_PORT
from db import store
from scrapers.tier1_easy import TIER1_SCRAPERS

# 收集所有可用爬虫（v1 加 TIER2_SCRAPERS 等）
ALL_SCRAPERS = {**TIER1_SCRAPERS}

log = logging.getLogger("runner")


def setup_logging(verbose=False):
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(os.path.join(LOG_DIR, "crawl.log"),
                                encoding="utf-8"),
        ],
    )
    os.makedirs(LOG_DIR, exist_ok=True)


def run_crawl(sources: list[str], max_per_source: int, skip_existing: bool = True):
    """同步跑一次爬虫"""
    started = time.time()
    log.info(f"▶▶ 启动爬取: sources={sources or 'all'}, max={max_per_source}, skip_existing={skip_existing}")

    target_classes = []
    for name, cls in ALL_SCRAPERS.items():
        if sources and name not in sources:
            continue
        src = store.get_source_by_name(name)
        if not src or not src["enabled"]:
            log.info(f"  跳过 {name} (disabled)")
            continue
        target_classes.append((name, cls))

    if not target_classes:
        log.error("没有可用爬虫")
        return

    total_found = 0
    total_new = 0
    for name, cls in target_classes:
        log.info(f"  → {name}")
        try:
            s = cls()
            found, new = s.crawl(max_articles=max_per_source, skip_existing=skip_existing)
            total_found += found
            total_new += new
        except Exception as e:
            log.exception(f"  ✗ {name} error: {e}")

    # 关 playwright
    try:
        from base_scraper import close_browser
        close_browser()
    except Exception:
        pass

    log.info(f"✓✓ 完成: 共 {len(target_classes)} 源, found={total_found}, new={total_new}, "
             f"耗时 {time.time()-started:.1f}s")


def serve_mode():
    """启动 Web + 后台持续爬虫（24/7）"""
    log.info("🌐 启动 Web 服务 + 后台持续爬虫")
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from web.app import app
        import uvicorn

        scheduler = BackgroundScheduler()

        # 每 5 分钟检查一次，随机决定是否跑各 tier
        scheduler.add_job(
            _random_crawl_cycle,
            "interval", minutes=5, id="random_crawl",
        )
        # 每天凌晨 3 点清理 3 天前旧数据
        scheduler.add_job(
            lambda: _cleanup_old_articles(3),
            "cron", hour=3, minute=17, id="cleanup",
        )
        # 定时导出：每天 8:00 和 0:00
        scheduler.add_job(
            _export_and_push, "cron", hour=0, minute=7, id="export_midnight",
        )
        scheduler.add_job(
            _export_and_push, "cron", hour=8, minute=7, id="export_morning",
        )
        scheduler.start()
        log.info("✓ 爬虫已启动：API 每 30m / HTTP 每 60m / PW 每 120m")
        log.info("✓ 定时导出：每天 0:00 / 8:00")

        log.info(f"💻 本地: http://localhost:{WEB_PORT}/matches")
        log.info(f"💻 本地: http://localhost:{WEB_PORT} (新闻列表)")

        # Web 用 8001 端口
        uvicorn.run(app, host="0.0.0.0", port=WEB_PORT)
    except KeyboardInterrupt:
        log.info("👋 退出")
    finally:
        try:
            scheduler.shutdown(wait=False)
        except Exception:
            pass


def _send_dingtalk_summary():
    """发送钉钉汇总"""
    try:
        from db import store
        from dingtalk import set_webhook, send_summary
        from config import DINGTALK_WEBHOOK_URL

        if not DINGTALK_WEBHOOK_URL:
            return
        set_webhook(DINGTALK_WEBHOOK_URL)

        stats = store.stats()
        sources = store.list_sources(enabled_only=False)
        source_detail = []
        for s in sources:
            cnt = s.get("article_count", 0) or 0
            if cnt > 0:
                source_detail.append({"name": s["display_name"] or s["name"], "new": cnt})
        source_detail.sort(key=lambda x: x["new"], reverse=True)

        matches = store.list_upcoming_matches(days=7, only_scheduled=False)
        total_m = len(matches)
        ok_m = sum(1 for m in matches if (m.get("article_count", 0) or 0) >= 16)
        low_m = total_m - ok_m

        send_summary({
            "total": stats.get("news_total", 0),
            "new_today": stats.get("news_today", 0),
            "sources": source_detail[:10],
        }, {
            "total": total_m,
            "达标": ok_m,
            "不足": low_m,
        })
    except Exception as e:
        log.warning(f"dingtalk summary error: {e}")


def _crawl_one_source(name: str, max_per: int):
    """爬单个源（给线程池调用）"""
    from scrapers.tier1_easy import TIER1_SCRAPERS
    from scrapers.playwright_sources import PLAYWRIGHT_SCRAPERS
    from scrapers.per_source import PER_SOURCE_SCRAPERS
    from web.crawl_status import crawl_status

    all_scrapers = {**TIER1_SCRAPERS, **PLAYWRIGHT_SCRAPERS, **PER_SOURCE_SCRAPERS}
    cls = all_scrapers.get(name)
    if not cls:
        return
    src = store.get_source_by_name(name)
    if not src or not src["enabled"]:
        return
    crawl_status.source_start(name)
    try:
        s = cls()
        found, new = s.crawl(max_articles=max_per, skip_existing=True)
        crawl_status.source_done(name, found, new, True)
        crawl_status.log_event(f"  ✓ {name}: 找到{found} 新增{new}")
    except Exception as e:
        crawl_status.source_done(name, 0, 0, False)
        crawl_status.log_event(f"  ✗ {name}: {e}", "error")


def _crawl_tier(tier_name: str, source_names: list[str], max_per: int):
    """并发爬指定 tier 的源（最多 3 个同时）"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from web.crawl_status import crawl_status

    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {pool.submit(_crawl_one_source, name, max_per): name for name in source_names}
        for fut in as_completed(futures):
            try:
                fut.result()
            except Exception as e:
                name = futures[fut]
                crawl_status.log_event(f"  ✗ {name}: {e}", "error")


def _random_crawl_cycle():
    """固定间隔 + 抖动的全源爬取（5 分钟检查一次）
    - API tier 每 30 分钟跑一次
    - HTTP tier 每 60 分钟跑一次
    - PW tier 每 120 分钟跑一次
    用 5 分钟时间桶判断，桶号 mod 间隔 == 偏移 时启动。
    每次启动前 sleep 0-90 秒随机抖动，避免规律性。
    """
    import random as _rand
    from match_tagger import tag_all_upcoming
    from web.crawl_status import crawl_status

    if crawl_status.is_running():
        return

    # 5 分钟时间桶（0, 1, 2, ... 一天 288 个）
    now_min = int(time.strftime("%M")) + int(time.strftime("%H")) * 60
    bucket = now_min // 5

    # 各 tier 的触发判断（错开桶避免同时启动）
    # API: 每 6 桶（30 min）触发一次，桶号 0
    # HTTP: 每 12 桶（60 min）触发一次，桶号 2
    # PW: 每 24 桶（120 min）触发一次，桶号 4
    triggers = {
        "API":  (bucket % 6 == 0,  ["dongqiudi", "espn_pw", "theanalyst_api", "bbc_sport", "skysports", "90min", "guardian"], 20),
        "HTTP": (bucket % 12 == 2, ["theanalyst", "flashscore", "goal", "fourfourtwo", "rotowire", "kickoff"], 15),
        "PW":   (bucket % 24 == 4, ["sofascore", "squawka", "fifa", "espn", "goal_pw"], 10),
    }

    any_ran = False
    for tier, (should_run, sources, max_per) in triggers.items():
        if not should_run:
            continue
        # 桶内抖动：启动前 sleep 0-90s，让请求时间不完全规律
        jitter = _rand.uniform(0, 90)
        time.sleep(jitter)
        if crawl_status.is_running():
            continue  # 抖动期间被其他周期抢占了
        crawl_status.start_crawl("")
        crawl_status.log_event(f"🔄 T{tier} 启动（抖动 {jitter:.0f}s）{time.strftime('%H:%M')} - {len(sources)} 源")
        _crawl_tier(tier, sources, max_per)
        any_ran = True

    if not any_ran:
        return  # 这个桶没有 tier 需要跑，直接退出

    # 重打标签
    try:
        crawl_status.log_event("  🏷 更新比赛标签…")
        tag_all_upcoming(days=7)
        matches = store.list_upcoming_matches(days=4, only_scheduled=False)
        ok = sum(1 for m in matches if (m.get("article_count", 0) or 0) >= 16)
        crawl_status.log_event(f"  比赛覆盖: {ok}/{len(matches)} 场达标")
    except: pass

    try:
        from base_scraper import close_browser
        close_browser()
    except: pass

    crawl_status.log_event(f"✓ 定时爬取完成 | 新增 {crawl_status.status().get('total_new', 0)} 篇")
    crawl_status.finish_crawl()


def _continuous_crawl_cycle():
    """后备：兜底全源爬取 + 重打标签"""
    from scrapers.match_harvester import run_all_sources
    from match_tagger import tag_all_upcoming
    try:
        run_all_sources(max_per_source=20)
        tag_all_upcoming(days=7)
    except Exception as e:
        log.exception(f"continuous cycle error: {e}")


def _cleanup_old_articles(days: int = 3):
    """清理 N 天前的旧文章"""
    import sqlite3
    cutoff = (f"{(time.strftime('%Y-%m-%d'))}" )  # today
    try:
        conn = sqlite3.connect(os.path.join(BASE_DIR, "db", "worldcup.db"))
        deleted = conn.execute(f"""
            DELETE FROM news WHERE published_at < date('now', '-{days} days')
               OR (published_at IS NULL AND crawled_at < datetime('now', '-{days} days'))
        """).rowcount
        conn.commit()
        conn.close()
        if deleted > 0:
            log.info(f"🧹 清理了 {deleted} 篇 {days} 天前旧文章")
    except Exception as e:
        log.warning(f"cleanup error: {e}")


def _load_env():
    """加载 .env 环境变量"""
    env_file = os.path.join(BASE_DIR, ".env")
    if os.path.exists(env_file):
        with open(env_file) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip())


def _export_and_push():
    """导出 4 天 Excel → push 到 GitHub → 钉钉通知"""
    import subprocess
    import io as _io

    log.info("📊 定时导出 Excel...")
    _load_env()

    try:
        # 生成 Excel
        from db import store
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        matches = store.list_upcoming_matches(days=4, only_scheduled=False, from_tomorrow=True)
        news = store.list_news(limit=5000, exclude_video=True)

        from collections import defaultdict
        from web.timezone_utils import beijing_day

        # 按北京时间日期分组
        day_matches = defaultdict(list)
        for m in matches:
            ka = m.get("kickoff_at") or ""
            day = beijing_day(ka) if ka else "未定"
            day_matches[day].append(m)

        export_dir = os.path.join(BASE_DIR, "data")
        os.makedirs(export_dir, exist_ok=True)
        token = os.environ.get("GITHUB_TOKEN", "")
        repo = os.environ.get("GITHUB_REPO", "")
        dingtalk_token = os.environ.get("DINGTALK_WEBHOOK_URL", "")

        files_pushed = []

        for day, day_ms in sorted(day_matches.items()):
            if day == "未定":
                continue

            wb = Workbook()
            for i, m in enumerate(day_ms):
                home = f"{m.get('home_cn','') or m.get('home_en','')}"
                away = f"{m.get('away_cn','') or m.get('away_en','')}"
                sheet_name = f"{home}vs{away}"[:31]
                cnt = m.get("article_count", 0) or 0

                if i == 0:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(sheet_name)

                ws.append(["标题", "来源", "时间", "URL", "摘要"])
                for cell in ws[1]:
                    cell.font = Font(bold=True)

                articles = store.list_match_news(m["id"], include_content=False)
                for n in articles:
                    ws.append([
                        n.get("title", ""),
                        n.get("source_name", ""),
                        (n.get("published_at") or "")[:19],
                        n.get("url", ""),
                        (n.get("summary") or "")[:150],
                    ])
                ws.column_dimensions['A'].width = 60
                ws.column_dimensions['B'].width = 18
                ws.column_dimensions['D'].width = 50

            day_short = day.replace("2026-", "")  # 0623
            day_display = day[5:].replace("-", "月") + "日"  # 6月23日
            update_date = time.strftime('%m%d')
            xlsx_name = f"worldcup_{day_short.replace('-','')}_{update_date}update.xlsx"
            xlsx_path = os.path.join(export_dir, xlsx_name)
            wb.save(xlsx_path)
            files_pushed.append((day_display, xlsx_name))
            log.info(f"✓ {day_display} → {xlsx_name} ({len(day_ms)} 场)")

        # Push
        if token and repo and files_pushed:
            gitkeep = os.path.join(export_dir, ".gitkeep")
            if not os.path.exists(gitkeep):
                with open(gitkeep, "w") as f: pass

            subprocess.run(["git", "-C", BASE_DIR, "add", "data/"], capture_output=True)
            subprocess.run(["git", "-C", BASE_DIR, "commit", "-m",
                           f"Auto export {time.strftime('%m%d')}"], capture_output=True)

            push_url = f"https://{token}@github.com/{repo}.git"
            r = subprocess.run(["git", "-C", BASE_DIR, "-c", "http.sslVerify=false", "push", push_url, "main"],
                             capture_output=True, text=True)
            if r.returncode == 0:
                log.info("✓ 已推送到 GitHub")
            else:
                log.warning(f"Git push failed: {r.stderr[:100]}")

        # 钉钉通知（纯文本，含每场比赛详情）
        if dingtalk_token and files_pushed:
            import requests as req
            from web.timezone_utils import to_beijing

            lines = ["⚽ 世界杯数据更新"]
            for day, day_ms in sorted(day_matches.items()):
                if day == "未定":
                    continue
                day_display = day[5:].replace("-", "月") + "日"
                day_short = day.replace("2026-", "").replace("-", "")
                update_date = time.strftime('%m%d')
                xlsx_name = f"worldcup_{day_short}_{update_date}update.xlsx"
                download_url = f"https://github.com/{repo}/raw/main/data/{xlsx_name}"

                lines.append(f"\n{day_display}")
                for m in day_ms:
                    home = f"{m.get('home_cn','') or m.get('home_en','')}"
                    away = f"{m.get('away_cn','') or m.get('away_en','')}"
                    bj_time = to_beijing(m.get("kickoff_at", ""), "%H:%M")
                    cnt = m.get("article_count", 0) or 0
                    lines.append(f"赛程: 北京时间{bj_time} {home}vs{away}")
                    lines.append(f"数据量: {cnt}条")
                lines.append(f"excel: {xlsx_name}")
                lines.append(f"下载: {download_url}")

            text = "\n".join(lines)
            try:
                req.post(dingtalk_token, json={"msgtype": "text", "text": {"content": text}}, timeout=10)
                log.info("✓ 钉钉通知已发送")
            except Exception as e:
                log.warning(f"钉钉通知失败: {e}")

    except Exception as e:
        log.exception(f"导出失败: {e}")


def main():
    p = argparse.ArgumentParser(description="世界杯新闻爬虫")
    p.add_argument("--quick", action="store_true", help="快速：每源 5 篇")
    p.add_argument("--backfill", action="store_true", help="存量：每源 20 篇")
    p.add_argument("--incremental", action="store_true", help="增量：每源 30 篇（跳过已有）")
    p.add_argument("--full", action="store_true", help="全量：每源 50 篇")
    p.add_argument("--source", type=str, default="", help="指定源，逗号分隔：fifa,espn")
    p.add_argument("--serve", action="store_true", help="启动 Web + 定时任务")
    p.add_argument("--init-db", action="store_true", help="初始化数据库")
    p.add_argument("--retag", action="store_true", help="重建所有新闻标签")
    # 比赛中心模式
    p.add_argument("--harvest-4d", action="store_true", help="为未来 4 天比赛每场挖 16 篇")
    p.add_argument("--incremental-3d", action="store_true", help="为未来 1-3 天比赛增量挖")
    p.add_argument("--harvest-match", type=int, metavar="ID", help="为指定比赛 ID 挖")
    p.add_argument("--refresh-schedule", action="store_true", help="增量刷新赛程（本地+openfootball+ESPN比分）")
    p.add_argument("--rebuild-schedule", action="store_true", help="清空 matches 表后重建（首次切换用）")
    p.add_argument("--retag-matches", action="store_true", help="重建所有比赛的新闻关联")
    p.add_argument("--verbose", "-v", action="store_true")
    args = p.parse_args()

    setup_logging(args.verbose)

    if args.init_db:
        from db.init_db import init_db
        init_db()
        return

    if args.retag:
        import tagger
        tagger.rebuild_all_tags()
        return

    if args.serve:
        serve_mode()
        return

    # 比赛中心模式
    if args.harvest_4d:
        from scrapers.match_harvester import harvest_upcoming
        harvest_upcoming(days=4, target=16)
        return
    if args.incremental_3d:
        from scrapers.match_harvester import harvest_upcoming
        harvest_upcoming(days=3, target=16)
        return
    if args.harvest_match:
        from scrapers.match_harvester import harvest_for_match
        m = store.get_match(args.harvest_match)
        if m:
            harvest_for_match(m, target=16)
        else:
            log.error(f"match {args.harvest_match} not found")
        return
    if args.refresh_schedule:
        from scrapers.schedule_builder import main as sched_main
        sched_main(rebuild=False)
        return
    if args.rebuild_schedule:
        from scrapers.schedule_builder import main as sched_main
        sched_main(rebuild=True)
        return
    if args.retag_matches:
        from match_tagger import tag_all_upcoming
        tag_all_upcoming(days=4)
        return

    # 确认 DB 已初始化
    if not os.path.exists(os.path.join(BASE_DIR, "db", "worldcup.db")):
        log.info("DB 不存在，先初始化…")
        from db.init_db import init_db
        init_db()

    sources = [s.strip() for s in args.source.split(",") if s.strip()] if args.source else []
    if args.quick:
        run_crawl(sources, max_per_source=5, skip_existing=True)
    elif args.backfill:
        run_crawl(sources, max_per_source=20, skip_existing=True)
    elif args.incremental:
        run_crawl(sources, max_per_source=30, skip_existing=True)
    elif args.full:
        run_crawl(sources, max_per_source=50, skip_existing=True)
    else:
        # 默认：跑 quick
        log.info("未指定模式，默认 --quick（5 篇/源）")
        run_crawl(sources, max_per_source=5, skip_existing=True)


if __name__ == "__main__":
    main()
