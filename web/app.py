"""
世界杯新闻爬虫 - FastAPI 前端
极简单页：列表 + 筛选 + 详情 + 触发爬取
"""
import os
import sys
import json
import threading
import time
from datetime import datetime, timezone, timedelta

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)

from fastapi import FastAPI, Request, Form, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from config import WEB_HOST, WEB_PORT
from db import store
from web.timezone_utils import to_beijing, beijing_day, beijing_relative_day

WEB_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(WEB_DIR, "templates")
STATIC_DIR = os.path.join(WEB_DIR, "static")

app = FastAPI(title="World Cup News Crawler")
templates = Jinja2Templates(directory=TEMPLATES_DIR)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

from web.crawl_status import crawl_status

# 爬虫锁
_crawl_lock = threading.Lock()


def _to_int(v):
    """表单提交空字符串时转 None"""
    if v is None or v == "" or v == "None":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


@app.get("/", response_class=HTMLResponse)
def index(request: Request,
          country: str = Query(None),
          player: str = Query(None),
          source: str = Query(None),
          q: str = Query(None),
          limit: int = Query(20)):
    country_id = _to_int(country)
    player_id = _to_int(player)
    # source 和 q 是字符串，空串转 None
    source_name = source if source else None
    query_str = q if q else None
    news = store.list_news(limit=limit, country_id=country_id, player_id=player_id,
                            source_name=source_name, query=query_str)
    total_count = store.count_news(country_id=country_id, player_id=player_id,
                                   source_name=source_name, query=query_str)
    countries = store.list_countries()
    # 球员：若选了国家，只显示该国的；否则全部
    players = store.list_players(country_id=country_id) if country_id else []
    sources = store.list_sources(enabled_only=True)
    return templates.TemplateResponse(request, "index.html", {
        "news": news,
        "countries": countries,
        "players": players,
        "sources": sources,
        "filters": {
            "country": country_id, "player": player_id,
            "source": source_name, "q": query_str,
        },
        "total": total_count, "shown": len(news),
        "crawl_status": crawl_status.status(),
    })


@app.get("/article/{news_id}", response_class=HTMLResponse)
def article(request: Request, news_id: int):
    n = store.get_news_by_id(news_id)
    if not n:
        return RedirectResponse(url="/", status_code=302)
    return templates.TemplateResponse(request, "article.html", {"n": n})


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    return templates.TemplateResponse(request, "dashboard.html", {"stats": store.stats()})


@app.post("/api/crawl/trigger")
def trigger_crawl(source: str = Form(None), quick: bool = Form(False)):
    """手动触发爬虫（后台异步跑，立即返回）"""
    if not _crawl_lock.acquire(blocking=False):
        return JSONResponse({"status": "busy", "msg": "已有爬虫在跑"}, status_code=409)
    t = threading.Thread(target=_run_crawl_async, args=(source, quick), daemon=True)
    t.start()
    return {"status": "started", "msg": "爬虫已启动"}


def _run_crawl_async(source_filter: str, quick: bool):
    try:
        crawl_status.start_crawl("")
        crawl_status.log_event(f"▶ 启动爬虫 (source={source_filter or 'all'}, quick={quick})")

        from scrapers.tier1_easy import TIER1_SCRAPERS
        from scrapers.playwright_sources import PLAYWRIGHT_SCRAPERS, _patch_fetch_playwright
        from scrapers.per_source import PER_SOURCE_SCRAPERS
        _patch_fetch_playwright()

        all_scrapers = {**TIER1_SCRAPERS, **PLAYWRIGHT_SCRAPERS, **PER_SOURCE_SCRAPERS}
        if source_filter:
            all_scrapers = {k: v for k, v in all_scrapers.items() if k == source_filter}

        total_found, total_new = 0, 0
        for name, cls in all_scrapers.items():
            src = store.get_source_by_name(name)
            if not src or not src["enabled"]:
                continue
            crawl_status.source_start(name)
            crawl_status.log_event(f"  → {name} 正在抓取…")
            try:
                s = cls()
                max_n = 5 if quick else 20
                found, new = s.crawl(max_articles=max_n)
                total_found += found
                total_new += new
                crawl_status.source_done(name, found, new, True)
                crawl_status.log_event(f"  ✓ {name}: 找到 {found} 新增 {new}")
            except Exception as e:
                crawl_status.source_done(name, 0, 0, False)
                crawl_status.log_event(f"  ✗ {name}: {e}", "error")

        try:
            from match_tagger import tag_all_upcoming
            crawl_status.log_event("  🏷 重打比赛标签…")
            tag_all_upcoming(days=7)
        except: pass

        try:
            from base_scraper import close_browser
            close_browser()
        except: pass

        crawl_status.log_event(f"✓ 全部完成: 共 {total_found} 篇, 新增 {total_new} 篇")
    except Exception as e:
        crawl_status.log_event(f"✗ 异常: {e}", "error")
    finally:
        crawl_status.finish_crawl()
        _crawl_lock.release()


@app.get("/api/crawl/status")
def api_crawl_status():
    return crawl_status.status()


@app.get("/api/crawl/stream")
async def api_crawl_stream(request: Request):
    """SSE 实时推送爬虫状态"""
    import asyncio
    from starlette.responses import StreamingResponse

    async def event_generator():
        q = crawl_status.subscribe()
        yield f"data: {json.dumps(crawl_status.status(), ensure_ascii=False)}\n\n"
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    msg = q.get(timeout=5)
                    yield f"data: {msg}\n\n"
                except:
                    yield f"data: {json.dumps(crawl_status.status(), ensure_ascii=False)}\n\n"
        finally:
            crawl_status.unsubscribe(q)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/api/news")
def api_news(country: int = None, player: int = None,
             source: str = None, q: str = None, limit: int = 50,
             offset: int = 0):
    """JSON API，给前端 fetch 用"""
    return {"news": store.list_news(limit=limit, offset=offset,
                                     country_id=country, player_id=player,
                                     source_name=source, query=q)}


# ============ 导出（用于导入到另一个系统） ============

def _build_export_row(n: dict) -> dict:
    """统一格式：标题 / 时间 / 来源 / 正文 / 国家 / 球员 / 语言 / URL"""
    return {
        "id": n.get("id"),
        "title": n.get("title") or "",
        "published_at": n.get("published_at") or "",
        "crawled_at": n.get("crawled_at") or "",
        "source_name": n.get("source_name") or "",
        "source_url": n.get("url") or "",
        "language": n.get("language") or "",
        "summary": n.get("summary") or "",
        "content": n.get("content") or "",
        "countries": (n.get("countries") or "").strip(","),
        "players": (n.get("players") or "").strip(","),
        "image_url": n.get("image_url") or "",
    }


@app.get("/api/export.json")
def export_json(country: int = None, player: int = None,
                source: str = None, q: str = None, limit: int = 1000):
    """导出 JSON：直接被另一个系统 fetch 拉取"""
    news = store.list_news(limit=limit, country_id=country, player_id=player,
                            source_name=source, query=q, include_content=True)
    return {
        "exported_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "total": total_count, "shown": len(news),
        "news": [_build_export_row(n) for n in news],
    }


@app.get("/api/export.csv")
def export_csv(country: int = None, player: int = None,
               source: str = None, q: str = None, limit: int = 1000):
    """导出 CSV：UTF-8 BOM（Excel 直接打开不乱码）"""
    import csv
    import io

    news = store.list_news(limit=limit, country_id=country, player_id=player,
                            source_name=source, query=q, include_content=True)
    rows = [_build_export_row(n) for n in news]

    buf = io.StringIO()
    buf.write("\ufeff")  # BOM
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=worldcup_news_{int(time.time())}.csv"},
    )


@app.get("/api/export.xlsx")
def export_xlsx(country: int = None, player: int = None,
                source: str = None, q: str = None, limit: int = 1000):
    """导出 Excel（openpyxl）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io as _io

    news = store.list_news(limit=limit, country_id=country, player_id=player,
                            source_name=source, query=q, include_content=True)
    rows = [_build_export_row(n) for n in news]

    wb = Workbook()
    ws = wb.active
    ws.title = "World Cup News"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        # 表头加粗
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([r[h] for h in headers])
        # 正文列宽加大 + 自动换行
        for col_idx, h in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 80 if h == "content" else 25
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=worldcup_news_{int(time.time())}.xlsx"},
    )


@app.get("/api/export.ndjson")
def export_ndjson(country: int = None, player: int = None,
                  source: str = None, q: str = None, limit: int = 1000):
    """NDJSON（每行一个 JSON）：流式导入另一个系统最方便"""
    news = store.list_news(limit=limit, country_id=country, player_id=player,
                            source_name=source, query=q, include_content=True)
    lines = "\n".join(json.dumps(_build_export_row(n), ensure_ascii=False) for n in news)
    return StreamingResponse(
        iter([lines]),
        media_type="application/x-ndjson",
    )


# ============ 比赛中心 API ============

def _match_summary(m: dict) -> dict:
    """比赛摘要（含进度）"""
    return {
        "id": m["id"],
        "home": m.get("home_cn") or m.get("home_en") or "?",
        "away": m.get("away_cn") or m.get("away_en") or "?",
        "home_en": m.get("home_en"),
        "away_en": m.get("away_en"),
        "kickoff_at": m.get("kickoff_at"),
        "stage": m.get("stage") or m.get("tournament_round"),
        "group": m.get("group_name"),
        "venue": m.get("venue"),
        "status": m.get("status"),
        "article_count": m.get("article_count", 0),
        "below_target": bool(m.get("below_target")),
        "below_16": (m.get("article_count", 0) or 0) < 16,
    }


@app.get("/api/matches")
def api_matches(days: int = 4, all: bool = False):
    """未来 N 天比赛列表 + 每场文章数"""
    if all:
        matches = store.list_all_matches(limit=200)
    else:
        matches = store.list_upcoming_matches(days=days, only_scheduled=False)
    return {
        "total": len(matches),
        "matches": [_match_summary(m) for m in matches],
    }


@app.get("/api/matches/{match_id}")
def api_match_detail(match_id: int):
    m = store.get_match(match_id)
    if not m:
        return JSONResponse({"error": "not found"}, status_code=404)
    news = store.list_match_news(match_id, include_content=False)
    return {
        "match": _match_summary(m),
        "article_count": len(news),
        "news": news,
    }


@app.get("/api/matches/{match_id}/news")
def api_match_news(match_id: int):
    """单场比赛的 16+ 篇文章 JSON（标准导出格式）"""
    news = store.list_match_news(match_id, include_content=True)
    return {
        "match_id": match_id,
        "total": total_count, "shown": len(news),
        "news": [_build_export_row(n) for n in news],
    }


@app.get("/api/matches/{match_id}/export.xlsx")
def api_match_export_xlsx(match_id: int):
    """导出某场比赛所有新闻为 Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io as _io

    news = store.list_match_news(match_id, include_content=True)
    rows = [_build_export_row(n) for n in news]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Match {match_id}"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([r[h] for h in headers])
        for col_idx, h in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 80 if h == "content" else 25
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=match_{match_id}_news.xlsx"},
    )


@app.post("/api/matches/{match_id}/harvest")
def api_harvest_match(match_id: int):
    """手动触发：为某场比赛挖 16 篇"""
    if not _crawl_lock.acquire(blocking=False):
        return JSONResponse({"status": "busy"}, status_code=409)
    m = store.get_match(match_id)
    if not m:
        _crawl_lock.release()
        return JSONResponse({"error": "match not found"}, status_code=404)
    t = threading.Thread(target=_run_match_harvest_async, args=(match_id,), daemon=True)
    t.start()
    return {"status": "started", "match_id": match_id,
            "match": f"{m.get('home_en')} vs {m.get('away_en')}"}


def _run_match_harvest_async(match_id: int):
    try:
        crawl_status.start_crawl(f"match_{match_id}")
        crawl_status.log_event(f"▶ 启动比赛 #{match_id} 挖掘")
        from scrapers.match_harvester import harvest_for_match
        m = store.get_match(match_id)
        r = harvest_for_match(m, target=16)
        crawl_status.log_event(f"✓ 完成 #{match_id}: 找到 {r.get('new_articles',0)}, "
                    f"关联 {r.get('total_linked',0)}, "
                    f"{'⚠ 不足 16' if r.get('below_target') else '✅ 达标'}")
    except Exception as e:
        crawl_status.log_event(f"✗ 异常: {e}", "error")
    finally:
        crawl_status.finish_crawl()
        _crawl_lock.release()


# ============ 比赛列表页 ============

@app.get("/matches", response_class=HTMLResponse)
def matches_page(request: Request, days: int = 4):
    matches = store.list_upcoming_matches(days=days, only_scheduled=False)
    # 按北京时间日期分组
    from collections import defaultdict, OrderedDict
    grouped = defaultdict(list)
    for m in matches:
        ka = m.get("kickoff_at") or ""
        day = beijing_day(ka) if ka else "未定"
        # 加北京时间显示
        s = _match_summary(m)
        s["kickoff_bj"] = to_beijing(ka, "%m-%d %H:%M") if ka else ""
        s["kickoff_time_bj"] = to_beijing(ka, "%H:%M") if ka else ""
        s["day_label"] = beijing_relative_day(ka)
        grouped[day].append(s)
    # 排序：4 天后优先（按"距今天数"降序），同日内按时间正序
    today_bj = to_beijing(datetime.now(timezone.utc).isoformat(), "%Y-%m-%d") if False else None
    from datetime import datetime
    now_bj_str = datetime.now(timezone(timedelta(hours=8))).strftime("%Y-%m-%d")
    def day_sort_key(d):
        if d == "未定": return (99, "")
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            now = datetime.strptime(now_bj_str, "%Y-%m-%d")
            diff = (dt - now).days
            # 4 天后优先（diff=4 排最前），所以用 -diff 排序
            return (-diff, d)
        except Exception:
            return (99, d)
    sorted_days = sorted(grouped.keys(), key=day_sort_key)
    # 给每天加 label
    day_labels = {d: beijing_relative_day(next(iter(grouped[d]))["kickoff_at"]) for d in grouped}
    # 统计 4 天后 vs 1-3 天
    bulk_count = sum(1 for m in matches if _days_until(m.get("kickoff_at")) == 4)
    incr_count = sum(1 for m in matches if 0 < _days_until(m.get("kickoff_at")) <= 3)

    return templates.TemplateResponse(request, "matches.html", {
        "grouped": grouped,
        "sorted_days": sorted_days,
        "day_labels": day_labels,
        "days": days,
        "total": len(matches),
        "bulk_count": bulk_count,
        "incr_count": incr_count,
        "crawl_status": crawl_status.status(),
    })


def _days_until(iso_str):
    """距今天数（按北京时间）"""
    if not iso_str: return -1
    try:
        from datetime import datetime
        bj = to_beijing(iso_str, "%Y-%m-%d")
        now_bj = datetime.now(timezone(timedelta(hours=8))).strftime("%Y-%m-%d")
        return (datetime.strptime(bj, "%Y-%m-%d") - datetime.strptime(now_bj, "%Y-%m-%d")).days
    except Exception:
        return -1


@app.get("/matches/{match_id}", response_class=HTMLResponse)
def match_detail_page(request: Request, match_id: int):
    m = store.get_match(match_id)
    if not m:
        return RedirectResponse(url="/matches", status_code=302)
    articles = store.list_match_news(match_id, include_content=True)
    # 给每条文章加北京时间
    for a in articles:
        a["bj_pub_time"] = to_beijing(a.get("published_at", ""), "%m-%d %H:%M")
    return templates.TemplateResponse(request, "match_detail.html", {
        "request": request,
        "m": _match_summary(m),
        "bj_kickoff": to_beijing(m.get("kickoff_at", ""), "%m-%d %H:%M"),
        "articles": articles,
        "crawl_status": crawl_status.status(),
    })


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=WEB_HOST, port=WEB_PORT)
