"""
生成合并的 4 天 Excel（含正文列），文件名固定 worldcup_4days_latest.xlsx
便于 viewer.html 读取 + 用户手动下载。

用法：
  python3 generate_combined_xlsx.py
"""
import os
import sys
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

from db import store
from web.timezone_utils import to_beijing, beijing_day
from match_keywords import generate_match_keywords, score_article_for_match
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _has_chinese(text: str) -> bool:
    return any('\u4e00' <= ch <= '\u9fff' for ch in (text or ""))


def _title_cn(title: str) -> str:
    """预留中文标题列：已有中文标题直接填入，英文标题留空等待翻译流程。"""
    return title if _has_chinese(title) else ""


def _adjust_relevance(relevance: dict, linked_count: int) -> dict:
    """被关联到太多比赛的文章通常是泛文章，除非它是直接对阵。"""
    out = dict(relevance)
    tier = out.get("tier")
    if linked_count >= 8 and tier != "A":
        out["label"] = "泛文章"
        out["score"] = min(float(out.get("score") or 0), 0.25)
        reason = out.get("reason") or ""
        out["reason"] = f"{reason}; 已关联{linked_count}场，疑似通用材料".strip("; ")
    elif not tier:
        out["label"] = "已过滤"
        out["score"] = 0.0
    return out


def main():
    ms = store.list_upcoming_matches(days=4, only_scheduled=False, from_tomorrow=True)
    day_matches = defaultdict(list)
    for m in ms:
        day = beijing_day(m.get('kickoff_at', '')) or '未定'
        day_matches[day].append(m)

    wb = Workbook()
    wb.remove(wb.active)

    total_articles = 0
    total_with_content = 0

    for day, day_ms in sorted(day_matches.items()):
        if day == '未定':
            continue
        # Sheet 名：6月24日
        sheet_name = f"{int(day[5:7])}月{int(day[8:10])}日"
        ws = wb.create_sheet(sheet_name[:31])

        headers = [
            "日期", "北京时间", "轮次", "比赛", "标题", "中文标题", "来源",
            "相关性", "相关性分", "命中原因", "关联场次",
            "URL", "发布时间", "摘要", "正文"
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(horizontal="center")
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 42
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 50
        ws.column_dimensions['M'].width = 16
        ws.column_dimensions['N'].width = 50
        ws.column_dimensions['O'].width = 100

        for m in day_ms:
            home = m.get('home_cn') or m.get('home_en', '')
            away = m.get('away_cn') or m.get('away_en', '')
            match_label = f"{home} vs {away}"
            bj_time = to_beijing(m.get('kickoff_at', ''), '%H:%M')
            round_label = m.get('tournament_round') or ''
            kws = generate_match_keywords(m)
            articles = store.list_match_news(m['id'], include_content=True)
            for n in articles:
                content = n.get('content') or n.get('summary') or ''
                if content:
                    total_with_content += 1
                total_articles += 1
                pub_at = (n.get('published_at') or '')[:16]  # YYYY-MM-DD HH:MM
                linked_count = int(n.get('linked_match_count') or 1)
                relevance = _adjust_relevance(
                    score_article_for_match(n.get('title') or '', content, kws),
                    linked_count,
                )
                ws.append([
                    day,
                    bj_time,
                    round_label,
                    match_label,
                    (n.get('title') or '')[:200],
                    _title_cn(n.get('title') or '')[:200],
                    n.get('source_name', ''),
                    relevance.get('label') or '',
                    round(float(relevance.get('score') or 0), 2),
                    relevance.get('reason') or '',
                    linked_count,
                    n.get('url', ''),
                    pub_at,
                    (n.get('summary') or '')[:150],
                    content[:30000],
                ])

    out = os.path.join(BASE_DIR, 'data', 'worldcup_4days_latest.xlsx')
    wb.save(out)
    size_kb = os.path.getsize(out) / 1024
    pct = (total_with_content * 100 // total_articles) if total_articles else 0
    print(f"✓ 已生成 {out}")
    print(f"  大小: {size_kb:.1f} KB")
    print(f"  文章: {total_articles} 篇 / 有正文 {total_with_content} 篇 ({pct}%)")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == '__main__':
    main()
